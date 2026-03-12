from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, View
from django.http import HttpResponse
from django.db.models import Q
from django.forms.models import model_to_dict

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from config.search_helpers import filter_by_text_query
from .models import Asset, Category
from .forms import AssetForm

from django.db.models import Sum
from staff.models import Employee
from service.models import ServiceTaskItem


# --- 1. СПИСОК МАЙНА (ListView) ---
class AssetListView(ListView):
    model = Asset
    template_name = 'inventory/asset_list.html'
    paginate_by = 30
    
    def get_queryset(self):
        search_query = self.request.GET.get('search', '').strip()
        category_id = self.request.GET.get('category', 'all')
        show_archived = self.request.GET.get('archived', 'false')

        # Фільтр по архіву
        assets_queryset = Asset.objects.filter(is_archived=(show_archived == 'true')).select_related('category')

        # Фільтр по категорії
        if category_id != 'all':
            assets_queryset = assets_queryset.filter(category_id=category_id).select_related('category')

        # Пошук через Python (збережено вашу логіку)
        if search_query:
            return filter_by_text_query(
                assets_queryset,
                search_query,
                lambda asset: f"{asset.name} {asset.inventory_number} {asset.barcode} {asset.location} {asset.get_account_display() if hasattr(asset, 'get_account_display') else str(asset.account)}"
            )
        return assets_queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # У вашому шаблоні очікується, що 'assets' - це об'єкт сторінки пагінації
        if context.get('page_obj'):
            context['assets'] = context['page_obj']
            
        context['categories'] = Category.objects.all()
        context['show_archived'] = self.request.GET.get('archived', 'false')
        context['search_query'] = self.request.GET.get('search', '')
        context['current_category'] = self.request.GET.get('category', 'all')
        
        qs = self.get_queryset()
        context['total_count'] = len(qs) if isinstance(qs, list) else qs.count()
        return context


# --- 2. СТВОРЕННЯ (CreateView) ---
class AssetCreateView(CreateView):
    model = Asset
    form_class = AssetForm
    template_name = 'inventory/asset_form.html'
    success_url = reverse_lazy('asset_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Додати майно'
        return context


# --- 3. РЕДАГУВАННЯ (UpdateView) ---
class AssetUpdateView(UpdateView):
    model = Asset
    form_class = AssetForm
    template_name = 'inventory/asset_form.html'
    success_url = reverse_lazy('asset_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Редагувати майно'
        return context


# --- 4. КЛОНУВАННЯ (Custom View) ---
class AssetCloneView(View):
    def get(self, request, pk):
        original_asset = get_object_or_404(Asset, pk=pk)
        initial_data = model_to_dict(original_asset)
        
        initial_data.pop('id', None)
        initial_data.pop('inventory_number', None)
        initial_data.pop('barcode', None)

        form = AssetForm(initial=initial_data)
        return render(request, 'inventory/asset_form.html', {
            'form': form,
            'title': f'Створення копії: {original_asset.name}'
        })

    def post(self, request, pk):
        form = AssetForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('asset_list')
        return render(request, 'inventory/asset_form.html', {'form': form})


# --- 5. АРХІВУВАННЯ (Custom View) ---
class AssetArchiveView(View):
    def post(self, request, pk):
        asset = get_object_or_404(Asset, pk=pk)
        reason = request.POST.get('archive_reason')
        date = request.POST.get('archive_date')

        if reason and date:
            asset.is_archived = True
            asset.archive_reason = reason
            asset.archive_date = date
            asset.save()

        return redirect('asset_list')


# --- 6. ЕКСПОРТ В EXCEL (Залишається без змін) ---
def export_assets_xlsx(request):
    search_query = request.GET.get('search', '').strip()
    category_id = request.GET.get('category', 'all')
    show_archived = request.GET.get('archived', 'false')

    # ДОДАНО select_related('category') для оптимізації
    if show_archived == 'true':
        assets = Asset.objects.filter(is_archived=True).select_related('category')
    else:
        assets = Asset.objects.filter(is_archived=False).select_related('category')

    if category_id and category_id != 'all':
        assets = assets.filter(category_id=category_id)

    if search_query:
        assets = assets.filter(
            Q(name__icontains=search_query) |
            Q(inventory_number__icontains=search_query) |
            Q(barcode__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(account__icontains=search_query) |
            Q(archive_reason__icontains=search_query)
        )

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventory_export.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Майно'

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align_center = Alignment(horizontal="center", vertical="center")
    cell_align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ['Баркод', 'Інвентарний №', 'Назва', 'Категорія', 'Розташування', 'Відповідальний', 'Рахунок', 'Дата придбання']
    if show_archived == 'true':
        headers.extend(['Причина архівування', 'Дата архівування'])

    ws.append(headers)

    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for asset in assets:
        cat_name = asset.category.name if asset.category else ''
        account_name = asset.get_account_display() if hasattr(asset, 'get_account_display') else asset.account
        row_data = [asset.barcode, asset.inventory_number, asset.name, cat_name, asset.location, asset.responsible_person, account_name, asset.purchase_date]
        if show_archived == 'true':
            row_data.extend([asset.archive_reason, asset.archive_date])

        ws.append(row_data)

        current_row = ws[ws.max_row]
        for cell in current_row:
            cell.border = thin_border
            if cell.column == 3:
                cell.alignment = cell_align_left
            else:
                cell.alignment = cell_align_center

    ws.column_dimensions['C'].width = 50
    for col in ws.columns:
        column_letter = col[0].column_letter
        if column_letter == 'C': continue

        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length + 4)
        if adjusted_width < 10: adjusted_width = 10
        if adjusted_width > 30: adjusted_width = 30
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(response)
    return response

def home_view(request):
    # 1. Майно
    assets_count = Asset.objects.filter(is_archived=False).count()

    # 2. Працівники
    employees_count = Employee.objects.count()

    # Очікують (Заявка ще не відправлена в сервіс)
    waiting_data = ServiceTaskItem.objects.filter(
        task__date_sent__isnull=True
    ).aggregate(total=Sum('quantity'))
    waiting_for_repair = waiting_data['total'] or 0

    # В ремонті (Заявка відправлена, картридж ще не повернувся)
    in_repair_data = ServiceTaskItem.objects.filter(
        task__date_sent__isnull=False,
        date_back_from_service__isnull=True
    ).aggregate(total=Sum('quantity'))
    in_repair_process = in_repair_data['total'] or 0

    # Готові на складі (Повернувся, але ще не виданий)
    ready_data = ServiceTaskItem.objects.filter(
        date_back_from_service__isnull=False,
        date_returned_to_user__isnull=True
    ).aggregate(total=Sum('quantity'))
    ready_on_stock = ready_data['total'] or 0

    return render(request, 'index.html', {
        'assets_count': assets_count,
        'waiting_for_repair': waiting_for_repair,
        'in_repair_process': in_repair_process,
        'ready_on_stock': ready_on_stock,
        'employees_count': employees_count
    })